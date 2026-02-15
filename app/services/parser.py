import io
import asyncio
import openpyxl
import zipfile
from typing import Dict, List
from app.models.domain import SheetMetadata, ColumnMetadata

class ExcelParser:
    # 1. Public async method to be called by Orchestrator
    async def parse_bytes(self, file_bytes: bytes) -> Dict[str, SheetMetadata]:
        """Parse Excel file asynchronously by offloading to a thread."""
        return await asyncio.to_thread(self._parse_sync, file_bytes)

    # 2. Public async method for data extraction
    async def extract_data(self, metadata: Dict[str, SheetMetadata], file_bytes: bytes) -> Dict[str, Dict[str, List]]:
        """Extract data asynchronously."""
        return await asyncio.to_thread(self._extract_data_sync, metadata, file_bytes)

    # 3. Internal synchronous logic (The Heavy Lifting)
    def _parse_sync(self, file_bytes: bytes) -> Dict[str, SheetMetadata]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except zipfile.BadZipFile:
            raise ValueError("Invalid file format. Please upload a valid .xlsx file.")
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {str(e)}")

        all_metadata = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.max_row <= 1:
                continue
            meta = self._extract_sheet_metadata(sheet)
            all_metadata[sheet_name] = meta
        return all_metadata

    def _extract_data_sync(self, metadata: Dict[str, SheetMetadata], file_bytes: bytes) -> Dict[str, Dict[str, List]]:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        data = {}
        for sheet_name, meta in metadata.items():
            sheet = wb[sheet_name]
            sheet_data = {}
            for col in meta.columns:
                col_values = []
                for row_idx in range(meta.header_row + 1, sheet.max_row + 1):
                    val = sheet.cell(row=row_idx, column=col.index).value
                    col_values.append(val)
                sheet_data[col.name] = col_values
            data[sheet_name] = sheet_data
        return data

    # Helper methods remain synchronous
    def _extract_sheet_metadata(self, sheet) -> SheetMetadata:
        header_row = self._detect_header_row(sheet)
        columns = []
        for col_idx in range(1, sheet.max_column + 1):
            header_val = sheet.cell(row=header_row, column=col_idx).value
            if not header_val:
                continue
            type_counts, sample_values = self._analyze_column(sheet, header_row, col_idx)
            dominant_type = max(type_counts, key=type_counts.get)
            if dominant_type == "empty": dominant_type = "string"
            columns.append(ColumnMetadata(
                name=str(header_val).strip(),
                index=col_idx,
                data_type=dominant_type,
                sample_values=sample_values[:5],
                non_empty_count=sum(v for k, v in type_counts.items() if k != "empty")
            ))
        return SheetMetadata(
            sheet_name=sheet.title,
            header_row=header_row,
            columns=columns,
            total_rows=sheet.max_row - header_row
        )

    def _detect_header_row(self, sheet) -> int:
        for row_idx in range(1, min(6, sheet.max_row + 1)):
            row_values = [sheet.cell(row=row_idx, column=c).value for c in range(1, sheet.max_column + 1)]
            non_empty = [v for v in row_values if v is not None]
            if len(non_empty) >= 2 and all(isinstance(v, str) for v in non_empty):
                return row_idx
        return 1

    def _analyze_column(self, sheet, header_row, col_idx):
        type_counts = {"string": 0, "number": 0, "date": 0, "bool": 0, "empty": 0}
        samples = []
        for row_idx in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is None:
                type_counts["empty"] += 1
            elif isinstance(val, bool):
                type_counts["bool"] += 1
            elif isinstance(val, (int, float)):
                type_counts["number"] += 1
                samples.append(val)
            elif isinstance(val, str):
                type_counts["string"] += 1
                samples.append(val)
            else:
                type_counts["string"] += 1
        return type_counts, samples





# import io
# import openpyxl
# import zipfile 
# from typing import Dict, List, Any
# from app.models.domain import SheetMetadata, ColumnMetadata

# class ExcelParser:
#     def parse_bytes(self, file_bytes: bytes) -> Dict[str, SheetMetadata]:
#         """Parse Excel file and extract metadata per sheet."""
#         try:
#             # openpyxl expects a seekable file-like object
#             wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
#         except zipfile.BadZipFile:
#             # This happens if the file is NOT a valid .xlsx (e.g., it's a CSV or .xls)
#             raise ValueError("Invalid file format. Please upload a valid .xlsx file (Excel 2007+). CSV or .xls files are not supported.")
#         except Exception as e:
#             raise ValueError(f"Failed to read Excel file: {str(e)}")

#         all_metadata = {}
        
#         for sheet_name in wb.sheetnames:
#             sheet = wb[sheet_name]
#             if sheet.max_row <= 1:
#                 continue
            
#             meta = self._extract_sheet_metadata(sheet)
#             all_metadata[sheet_name] = meta
            
#         return all_metadata

#     def extract_data(self, metadata: Dict[str, SheetMetadata], file_bytes: bytes) -> Dict[str, Dict[str, List]]:
#         """Extract raw data columns from file based on metadata."""
#         wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
#         data = {}
        
#         for sheet_name, meta in metadata.items():
#             sheet = wb[sheet_name]
#             sheet_data = {}
            
#             for col in meta.columns:
#                 col_values = []
#                 for row_idx in range(meta.header_row + 1, sheet.max_row + 1):
#                     val = sheet.cell(row=row_idx, column=col.index).value
#                     col_values.append(val)
#                 sheet_data[col.name] = col_values
            
#             data[sheet_name] = sheet_data
#         return data

#     def _extract_sheet_metadata(self, sheet) -> SheetMetadata:
#         header_row = self._detect_header_row(sheet)
#         columns = []
        
#         for col_idx in range(1, sheet.max_column + 1):
#             header_val = sheet.cell(row=header_row, column=col_idx).value
#             if not header_val:
#                 continue
            
#             # Analyze column types
#             type_counts, sample_values = self._analyze_column(sheet, header_row, col_idx)
#             dominant_type = max(type_counts, key=type_counts.get)
#             if dominant_type == "empty": dominant_type = "string"
            
#             columns.append(ColumnMetadata(
#                 name=str(header_val).strip(),
#                 index=col_idx,
#                 data_type=dominant_type,
#                 sample_values=sample_values[:5],
#                 non_empty_count=sum(v for k, v in type_counts.items() if k != "empty")
#             ))
            
#         return SheetMetadata(
#             sheet_name=sheet.title,
#             header_row=header_row,
#             columns=columns,
#             total_rows=sheet.max_row - header_row
#         )

#     def _detect_header_row(self, sheet) -> int:
#         for row_idx in range(1, min(6, sheet.max_row + 1)):
#             row_values = [sheet.cell(row=row_idx, column=c).value for c in range(1, sheet.max_column + 1)]
#             non_empty = [v for v in row_values if v is not None]
#             if len(non_empty) >= 2 and all(isinstance(v, str) for v in non_empty):
#                 return row_idx
#         return 1

#     def _analyze_column(self, sheet, header_row, col_idx):
#         type_counts = {"string": 0, "number": 0, "date": 0, "bool": 0, "empty": 0}
#         samples = []
        
#         for row_idx in range(header_row + 1, min(header_row + 50, sheet.max_row + 1)):
#             val = sheet.cell(row=row_idx, column=col_idx).value
#             if val is None:
#                 type_counts["empty"] += 1
#             elif isinstance(val, bool):
#                 type_counts["bool"] += 1
#             elif isinstance(val, (int, float)):
#                 type_counts["number"] += 1
#                 samples.append(val)
#             elif isinstance(val, str):
#                 type_counts["string"] += 1
#                 samples.append(val)
#             else:
#                 type_counts["string"] += 1
                
#         return type_counts, samples